"""
Base output formatter for extensible multi-format output generation.

This module provides a comprehensive base class for output formatting that supports:
- Multiple output formats (table, json, csv, yaml, markdown, html)
- Human-readable and machine-readable formats
- Extensible architecture for custom formats
- Consistent interface across all tools
"""

import csv
import io
import json
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Protocol


class OutputFormat:
    """Enumeration of supported output formats."""

    TABLE = "table"
    JSON = "json"
    CSV = "csv"
    YAML = "yaml"
    MARKDOWN = "markdown"
    HTML = "html"
    XML = "xml"
    TSV = "tsv"
    EXCEL = "excel"  # Requires additional dependency


class FormatterProtocol(Protocol):
    """Protocol for output formatters."""

    def format(self, data: dict[str, Any], **kwargs) -> str:
        """Format data to string output."""
        ...


class BaseOutputFormatter(ABC):
    """
    Abstract base class for all output formatters.

    Provides a consistent interface and common functionality for formatting
    data into various output formats.
    """

    def __init__(self):
        """Initialize the formatter with format handlers."""
        self._format_handlers = {
            OutputFormat.TABLE: self._format_table,
            OutputFormat.JSON: self._format_json,
            OutputFormat.CSV: self._format_csv,
            OutputFormat.YAML: self._format_yaml,
            OutputFormat.MARKDOWN: self._format_markdown,
            OutputFormat.HTML: self._format_html,
            OutputFormat.XML: self._format_xml,
            OutputFormat.TSV: self._format_tsv,
        }

    def format(
        self, data: dict[str, Any], format_type: str = OutputFormat.TABLE, **kwargs
    ) -> str:
        """
        Format data according to the specified format type.

        Args:
            data: Data to format
            format_type: Output format type
            **kwargs: Additional format-specific options

        Returns:
            Formatted string output
        """
        handler = self._format_handlers.get(format_type)
        if not handler:
            raise ValueError(f"Unsupported format type: {format_type}")

        return handler(data, **kwargs)

    def save(
        self,
        data: dict[str, Any],
        output_path: str | Path,
        format_type: str = OutputFormat.JSON,
        **kwargs,
    ) -> None:
        """
        Save formatted data to a file.

        Args:
            data: Data to save
            output_path: Path to save the file
            format_type: Output format type
            **kwargs: Additional format-specific options
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        content = self.format(data, format_type, **kwargs)

        # Special handling for Excel format (would need openpyxl)
        if format_type == OutputFormat.EXCEL:
            self._save_excel(data, output_path, **kwargs)
        else:
            output_path.write_text(content, encoding="utf-8")

    @abstractmethod
    def _format_table(self, data: dict[str, Any], **kwargs) -> str:
        """Format data as a human-readable table."""
        pass

    def _format_json(self, data: dict[str, Any], **kwargs) -> str:
        """Format data as JSON."""
        indent = kwargs.get("indent", 2)
        sort_keys = kwargs.get("sort_keys", True)
        return json.dumps(data, indent=indent, sort_keys=sort_keys, default=str)

    @abstractmethod
    def _format_csv(self, data: dict[str, Any], **kwargs) -> str:
        """Format data as CSV."""
        pass

    def _format_yaml(self, data: dict[str, Any], **kwargs) -> str:
        """Format data as YAML."""
        try:
            import yaml

            return yaml.dump(data, default_flow_style=False, sort_keys=True)
        except ImportError as e:
            raise ImportError("PyYAML is required for YAML output format") from e

    def _format_markdown(self, data: dict[str, Any], **kwargs) -> str:
        """Format data as Markdown."""
        # Default implementation - tools should override for better formatting
        lines = [f"# {data.get('title', 'Output')}"]
        lines.append("")

        # Add metadata section
        metadata = data.get("metadata", {})
        if metadata:
            lines.append("## Metadata")
            lines.append("")
            for key, value in metadata.items():
                lines.append(f"- **{key}**: {value}")
            lines.append("")

        # Add data section
        lines.append("## Data")
        lines.append("")
        lines.append("```json")
        lines.append(json.dumps(data.get("data", data), indent=2, default=str))
        lines.append("```")

        return "\n".join(lines)

    def _format_html(self, data: dict[str, Any], **kwargs) -> str:
        """Format data as HTML."""
        # Basic HTML template - tools should override for better formatting
        title = data.get("title", "Output")
        content = self._format_markdown(data, **kwargs)

        # Convert markdown to HTML (simplified)
        content = content.replace("# ", "<h1>").replace("\n\n", "</h1>\n\n")
        content = content.replace("## ", "<h2>").replace("\n\n", "</h2>\n\n")
        content = content.replace("```json\n", "<pre><code>")
        content = content.replace("\n```", "</code></pre>")

        return f"""<!DOCTYPE html>
<html>
<head>
    <title>{title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        pre {{ background: #f4f4f4; padding: 10px; overflow-x: auto; }}
        code {{ font-family: monospace; }}
    </style>
</head>
<body>
    {content}
</body>
</html>"""

    def _format_xml(self, data: dict[str, Any], **kwargs) -> str:
        """Format data as XML."""
        try:
            import xml.etree.ElementTree as ET
            from xml.dom import minidom

            root = ET.Element("root")
            self._dict_to_xml(data, root)

            # Pretty print
            rough_string = ET.tostring(root, "unicode")
            reparsed = minidom.parseString(rough_string)
            return reparsed.toprettyxml(indent="  ")
        except ImportError as e:
            raise ImportError("xml module is required for XML output format") from e

    def _dict_to_xml(self, d: dict | list | Any, parent: Any) -> None:
        """Convert dictionary to XML elements recursively."""
        if isinstance(d, dict):
            for key, value in d.items():
                # Sanitize key for XML
                key = str(key).replace(" ", "_").replace("-", "_")
                child = parent.find(key)
                if child is None:
                    child = parent.makeelement(key, {})
                    parent.append(child)
                self._dict_to_xml(value, child)
        elif isinstance(d, list):
            for item in d:
                item_elem = parent.makeelement("item", {})
                parent.append(item_elem)
                self._dict_to_xml(item, item_elem)
        else:
            parent.text = str(d)

    def _format_tsv(self, data: dict[str, Any], **kwargs) -> str:
        """Format data as TSV (Tab-Separated Values)."""
        # Similar to CSV but with tabs
        csv_content = self._format_csv(data, **kwargs)
        # Convert CSV to TSV
        output = io.StringIO()
        reader = csv.reader(io.StringIO(csv_content))
        writer = csv.writer(output, delimiter="\t")
        for row in reader:
            writer.writerow(row)
        return output.getvalue()

    def _save_excel(self, data: dict[str, Any], output_path: Path, **kwargs) -> None:
        """Save data as Excel file (requires openpyxl)."""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            if ws:
                ws.title = data.get("title", "Data")[:31]  # Excel sheet name limit

            # Convert to CSV first, then to Excel
            csv_content = self._format_csv(data, **kwargs)
            reader = csv.reader(io.StringIO(csv_content))

            if ws:
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    if column[0].column is not None:
                        column_letter = get_column_letter(column[0].column)
                    else:
                        continue
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (AttributeError, TypeError):
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(output_path)
        except ImportError as e:
            raise ImportError("openpyxl is required for Excel output format") from e


class TableFormatter:
    """Helper class for creating formatted tables."""

    @staticmethod
    def create_table(
        headers: list[str],
        rows: list[list[str]],
        column_widths: list[int] | None = None,
        alignment: str = "left",
    ) -> str:
        """
        Create a formatted text table.

        Args:
            headers: Column headers
            rows: Data rows
            column_widths: Optional fixed column widths
            alignment: Text alignment (left, center, right)

        Returns:
            Formatted table as string
        """
        if not column_widths:
            # Calculate column widths
            column_widths = [len(h) for h in headers]
            for row in rows:
                for i, cell in enumerate(row):
                    column_widths[i] = max(column_widths[i], len(str(cell)))

        # Create format strings
        if alignment == "center":
            formats = [f"{{:^{w}}}" for w in column_widths]
        elif alignment == "right":
            formats = [f"{{:>{w}}}" for w in column_widths]
        else:  # left
            formats = [f"{{:<{w}}}" for w in column_widths]

        # Build table
        lines = []

        # Header
        header_row = " | ".join(
            fmt.format(h) for fmt, h in zip(formats, headers, strict=False)
        )
        lines.append(header_row)
        lines.append("-" * len(header_row))

        # Data rows
        for row in rows:
            row_str = " | ".join(
                fmt.format(str(cell)) for fmt, cell in zip(formats, row, strict=False)
            )
            lines.append(row_str)

        return "\n".join(lines)
